# backend/file_extractors.py
from __future__ import annotations

from io import BytesIO
from typing import Optional


def _safe_decode_text(b: bytes) -> str:
    for enc in ("utf-8", "utf-16", "latin-1"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return ""


def _extract_pdf(pdf_bytes: bytes) -> str:
    # Try PyMuPDF, then pdfplumber, then pypdf
    errors = []
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        parts = [(p.get_text("text") or "") for p in doc]
        doc.close()
        txt = "\n".join(parts).strip()
        if len(txt) > 20:
            return txt
    except Exception as e:
        errors.append(f"pymupdf: {e}")

    try:
        import pdfplumber
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            parts = [(p.extract_text() or "") for p in pdf.pages]
        txt = "\n".join(parts).strip()
        if len(txt) > 20:
            return txt
    except Exception as e:
        errors.append(f"pdfplumber: {e}")

    try:
        from pypdf import PdfReader
        reader = PdfReader(BytesIO(pdf_bytes))
        parts = [(p.extract_text() or "") for p in reader.pages]
        txt = "\n".join(parts).strip()
        if len(txt) > 20:
            return txt
    except Exception as e:
        errors.append(f"pypdf: {e}")

    return ""


def _extract_docx(docx_bytes: bytes) -> str:
    try:
        from docx import Document  # python-docx
    except Exception:
        return ""
    try:
        doc = Document(BytesIO(docx_bytes))
        parts = [p.text for p in doc.paragraphs if p.text and p.text.strip()]
        return "\n".join(parts).strip()
    except Exception:
        return ""


def _extract_xlsx(xlsx_bytes: bytes) -> str:
    # Extract text-ish cells to a readable block
    try:
        from openpyxl import load_workbook
    except Exception:
        return ""
    try:
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
        out = []
        for ws in wb.worksheets:
            out.append(f"\n=== SHEET: {ws.title} ===")
            # limit rows/cols to avoid huge dumps
            max_r = min(ws.max_row or 0, 200)
            max_c = min(ws.max_column or 0, 30)
            for r in range(1, max_r + 1):
                row_vals = []
                for c in range(1, max_c + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is None:
                        row_vals.append("")
                    else:
                        row_vals.append(str(v))
                if any(x.strip() for x in row_vals):
                    out.append(" | ".join(row_vals))
        txt = "\n".join(out).strip()
        return txt
    except Exception:
        return ""


def extract_text_from_any(filename: str, content_type: str, data: bytes) -> str:
    name = (filename or "").lower().strip()
    ctype = (content_type or "").lower().strip()

    # PDFs
    if ctype == "application/pdf" or name.endswith(".pdf"):
        return _extract_pdf(data)

    # DOCX
    if ctype in ("application/vnd.openxmlformats-officedocument.wordprocessingml.document",) or name.endswith(".docx"):
        return _extract_docx(data)

    # XLSX
    if ctype in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",) or name.endswith(".xlsx"):
        return _extract_xlsx(data)

    # Text-ish (txt/md/csv/json)
    if ctype.startswith("text/") or name.endswith((".txt", ".md", ".csv", ".json", ".log")):
        return _safe_decode_text(data).strip()

    return ""
